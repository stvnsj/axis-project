import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
import annexUtils
from openpyxl import load_workbook


OFFSET   = 38
PAGEBREAKS = []

def generate (input_file='anexos/anteproyecto/annex1.xlsx',output_file="test2.xlsx") :
    
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet("2.903.3.F (RRP)")
    writer = Writer(workbook,worksheet)
    
    COL_WIDTHS = [
        0.10, #A 
        0.30, #B
        0.75, #C
        0.25, #D
        0.35, #E
        0.18, #F
        0.19, #G
        0.19, #H
        0.25, #I
        0.22, #J
        0.30, #K
        0.25, #L
        0.15, #M
        0.35, #N
        0.15, #O
        0.20, #P
        0.20, #Q
        0.33, #R
        0.25, #S
        0.15, #T
        0.25, #U
        0.15, #V
        0.33, #W
        0.19, #X
        0.46, #Y
        0.25, #Z
        0.10, #AA
    ]
    
    ROW_DICT = {
        0 :0.1,  
        6 :0.1,  
        7 :0.18, 
        8 :0.18, 
        9 :0.19, 
        10:0.1,
        12:0.12,
        13:0.12
        
    }
    
    ROW_DICT[1 - 1]  = 0.1
    ROW_DICT[7 - 1]  = 0.1
    ROW_DICT[11 - 1] = 0.1 
    
    #worksheet.autofit()
    annexUtils.set_column(worksheet,COL_WIDTHS)
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
    
    # FIXED CONTENT
    writer.merge(f"B2:F6","",Format.BORDER)
    writer.merge(f"G2:Z3","PUNTOS DE LA RED DE REFERENCIA PRINCIPAL",
                 {"top":1, "right":1,"font_size":12,"bold":True,"align":"center","valign":"vcenter"})
    
    writer.merge(f"G4:Z6","FORMULARIO N° 2.903.3.F",
                 {"bottom":1,"right":1, "font_size":12,"bold":True,"align":"center","valign":"vcenter"})
    
    writer.write(f"AA2","",Format.LEFT)

    writer.merge(f"B7:Z7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"AA8:AA12","",{"left":1})
    writer.merge(f"B13:Z13", "",{"top":1})

    writer.merge(f"B8:D8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B9:D9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B10:D10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B12:D12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"U12:Z12",f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
  
    

    
    
    wb = load_workbook(input_file)
    ws = wb.active
    VALUE = ws['C22'].value
    FRST_ROW = 0
    LST_ROW  = 0
    
    # Iterate through merged cell ranges
    for merged_cell_range in ws.merged_cells.ranges:
        # Check if the merged region is in column B
        if merged_cell_range.min_col == 2 and merged_cell_range.max_col == 2:  # Column B has index 2
            # Get the first cell of the merged region
            first_cell = ws.cell(row=merged_cell_range.min_row, column=2).value
            print(first_cell)
            # Check if the first cell contains the text "RBP"
            if first_cell == "RRP":
                # Get the first and last rows of the merged region
                FRST_ROW = merged_cell_range.min_row
                LST_ROW  = merged_cell_range.max_row
                first_row = merged_cell_range.min_row
                last_row = merged_cell_range.max_row
                # print(f"Merged region with 'RBP' starts at row {first_row} and ends at row {last_row}")
                break
 
    scanner = annexUtils.Scanner(ws)
    
    
    # LOOP THE FOLLOWING CELLS
    CORRECTION = 0
    for i,r in enumerate(range(FRST_ROW,LST_ROW+1)):
        
        CONST = i * OFFSET + CORRECTION
        
        
        CELL_nombre = ws[f'{scanner.PRO.column_letter}{r}'].value # DONE
        
        CELL_f      = ws[f'{scanner.GEO_S.column_letter}{r}'].value
        CELL_l      = ws[f'{scanner.GEO_W.column_letter}{r}'].value
        CELL_h      = ws[f'{scanner.ELIP.column_letter}{r}'].value
        
        CELL_X      = ws[f'{scanner.GEO_X.column_letter}{r}'].value
        CELL_Y      = ws[f'{scanner.GEO_Y.column_letter}{r}'].value
        CELL_Z      = ws[f'{scanner.GEO_Z.column_letter}{r}'].value
        
        CELL_N      = ws[f'{scanner.UTM_N.column_letter}{r}'].value
        CELL_E      = ws[f'{scanner.UTM_E.column_letter}{r}'].value
        
        CELL_altura = ws[f'{scanner.COTA_ORTO.column_letter}{r}'].value
        CELL_cota   = ws[f'{scanner.COTA_GEO.column_letter}{r}'].value
        
        CELL_NL     = ws[f'K{r}'].value
        CELL_EL     = ws[f'L{r}'].value
        
        CELL_MCL    = ws['H9'].value
        CELL_Ko     = ws['H12'].value
        
        
        
        writer.write(f"B{15 + CONST}","Identificación del Punto",Format.BOLD, Format.ITALIC, Format.SIZE(10))
        writer.write(f"F{15 + CONST}", "Nombre:", Format.SIZE(10))
        writer.write(f"N{15 + CONST}", "Dm. Ref.:",Format.SIZE(10))
        
        writer.merge(
            f"I{15 + CONST}:L{15 + CONST}",
            CELL_nombre,
            Format.SIZE(11), Format.BOTTOM, Format.CENTER
        )
        
        writer.write(
            f"C{21 + CONST}",
            CELL_f,
        )
        
        writer.write(
            f"C{22 + CONST}",
            CELL_l,
        )
        
        writer.write(
            f"C{23 + CONST}",
            CELL_h,
            Format.CENTER
        )
        
        writer.merge(
            f"H{21 + CONST}:K{21 + CONST}",
            CELL_X,
            Format.RIGHT
        )
        
        writer.merge(
            f"H{22 + CONST}:K{22 + CONST}",
            CELL_Y,
            Format.RIGHT
        )
        
        writer.merge(
            f"H{23 + CONST}:K{23 + CONST}",
            CELL_Z,
            Format.RIGHT
        )
        
        writer.merge(f'P{20 + CONST}:R{20 + CONST}',scanner.ZONA, Format.CENTER)
        writer.merge(f'P{21 + CONST}:R{21 + CONST}',scanner.MC, Format.CENTER)
        writer.merge(f"O{22 + CONST}:R{22 + CONST}",CELL_N,Format.RIGHT)
        writer.merge(f"O{23 + CONST}:R{23 + CONST}",CELL_E,Format.RIGHT)
        
        
        writer.merge(f"Q{15 + CONST}:T{15 + CONST}", "",Format.CENTER,Format.SIZE(10),Format.BOTTOM)
        writer.write(f"W{15 + CONST}","FECHA:",Format.SIZE(10))
        writer.merge(f"Y{15 + CONST}:Z{15 + CONST}", annexUtils.curr_date(1),Format.BOTTOM,Format.CENTER,Format.SIZE(10))
        
        writer.merge(f"B{17 + CONST}:Z{17 + CONST}", "Coordenadas",Format.SIZE(11),Format.CENTER)
        writer.merge(f"B{18 + CONST}:F{18 + CONST}", "Geodésicas",Format.SIZE(11),Format.CENTER)
        writer.merge(f"B{19 + CONST}:F{19 + CONST}", "Ref. SIRGAS", Format.SIZE(11),Format.CENTER)
        
        writer.merge(f"G{18 + CONST}:L{19 + CONST}", "Geocéntricas",Format.SIZE(11), Format.CENTER, Format.VCENTER )
        writer.merge(f"N{18 + CONST}:S{19 + CONST}", "UTM",Format.SIZE(11),Format.CENTER,Format.VCENTER)
        
        
        writer.write(f"B{21 + CONST}","f:",Format.SIZE(11),Format.LEFT)
        writer.write(f"B{22 + CONST}","l:",Format.SIZE(11),Format.LEFT)
        writer.write(f"B{23 + CONST}","h:",Format.SIZE(10),Format.LEFT)
        
        writer.write(f"G{21 + CONST}","X:",Format.SIZE(11),Format.LEFT)
        writer.write(f"G{22 + CONST}","Y:",Format.SIZE(11),Format.LEFT)
        writer.write(f"G{23 + CONST}","Z:",Format.SIZE(11),Format.LEFT)
        
        
        writer.write(f"L{21 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"L{22 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"L{23 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"D{23 + CONST}","m",Format.SIZE(11),Format.LEFT)

        writer.write(f"S{22 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"S{23 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"Z{22 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"Z{23 + CONST}","m",Format.SIZE(11),Format.LEFT)
        
        
        
        writer.write(f"N{20 + CONST}","Huso:",Format.SIZE(11),Format.LEFT)
        writer.write(f"N{21 + CONST}","MC:",Format.SIZE(11),Format.LEFT)
        writer.write(f"N{22 + CONST}","N:",Format.SIZE(11),Format.LEFT)
        writer.write(f"N{23 + CONST}","E:",Format.SIZE(11),Format.LEFT)
        
        
        #############
        # ITERATION #
        #############
        
        PTL_OFFSET = 0
        
        for k in range(len(scanner.PTL_N)):
            
            LTM_N = ws[f'{scanner.PTL_N[k].column_letter}{r}'].value
            LTM_E = ws[f'{scanner.PTL_E[k].column_letter}{r}'].value
            MCL   = scanner.MERIDIANO_CENTRAL[k].value
            FACTOR = scanner.FACTOR_ESCALA[k].value
            
            try:
                float(LTM_N)
                
            except:
                continue
            
            writer.merge(f"U{18 + CONST + PTL_OFFSET }:Z{19 + CONST + PTL_OFFSET}", f"PTL-{k+1}",Format.SIZE(11),Format.CENTER,Format.VCENTER)
            writer.write(f"U{20 + CONST + PTL_OFFSET }","MCL:",Format.SIZE(11),Format.LEFT)
            writer.write(f"U{21 + CONST + PTL_OFFSET}","Ko:",Format.SIZE(11),Format.LEFT)
            writer.write(f"U{22 + CONST + PTL_OFFSET}","NL:",Format.SIZE(11),Format.LEFT)
            writer.write(f"U{23 + CONST + PTL_OFFSET}","EL:",Format.SIZE(11),Format.LEFT)
            writer.merge(f"W{20 + CONST + PTL_OFFSET}:Z{20 + CONST + PTL_OFFSET}", MCL, Format.LEFT )
            writer.merge(f"W{21 + CONST + PTL_OFFSET}:Z{21 + CONST + PTL_OFFSET}", FACTOR, Format.LEFT )
            writer.merge(f"V{22 + CONST + PTL_OFFSET}:Y{22 + CONST + PTL_OFFSET}",LTM_N,Format.RIGHT )
            writer.merge(f"V{23 + CONST + PTL_OFFSET}:Y{23 + CONST + PTL_OFFSET}",LTM_E,Format.RIGHT )
            PTL_OFFSET += 6
            
        
        CORRECTION += PTL_OFFSET - 6
        CONST += PTL_OFFSET - 6
        
        #-------------------------------------------------------------------------------
        ###############
        # SUB 24 ROWS #
        ###############
        writer.write(f"F{25 + CONST}","Altura (n.m.m. modelada):",Format.SIZE(11))
        writer.write(f"S{25 + CONST}","Cota (nivelada):",Format.SIZE(11),Format.BOLD)
        writer.write(f"Q{25 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"Z{25 + CONST}","m",Format.SIZE(11),Format.LEFT)
        writer.merge(f"N{25 + CONST}:P{25 + CONST}",CELL_altura, Format.RIGHT)
        writer.merge(f"X{25 + CONST}:Y{25 + CONST}",CELL_cota,Format.RIGHT)
        
        
        
        writer.merge(f"B{27 + CONST}:K{39 + CONST}","Fotografía\nPanorámica",Format.BORDER,Format.CENTER,Format.VCENTER)
        writer.merge(f"M{27 + CONST}:Z{39 + CONST}","Vista\nAérea",Format.BORDER,Format.CENTER,Format.VCENTER)
        writer.merge(f"B{41 + CONST}:F{48 + CONST}","Fotografía\nDetalle",Format.BORDER,Format.CENTER,Format.VCENTER)
        
        
        writer.merge(f"H{41 + CONST}:Z{41 + CONST}","Descripción",Format.BOTTOM,Format.LEFT,Format.SIZE(10))
        
        writer.write(f'I{43 + CONST}', "Materialidad:",Format.SIZE(9))
        writer.write(f'I{44 + CONST}', "Dimensiones:",Format.SIZE(9))
        writer.write(f'I{45 + CONST}', "Distancia a :",Format.SIZE(9))
        
        writer.write(f'L{43 + CONST}', "MONOLITO DE HORMIGÓN, PINTADO DE AMARILLO", Format.SIZE(9))
        writer.write(f'L{44 + CONST}', "30,0 cm x 30,0 cm x 50,0 cm", Format.SIZE(9))
        
        writer.merge(f"G{42 + CONST}:G{48 + CONST}","",Format.BRIGHT)
        writer.merge(f"AA{42 + CONST}:AA{48 + CONST}","",Format.BLEFT)
        writer.merge(f"H{49 + CONST}:Z{49 + CONST}","",Format.TOP)
        writer.merge(f"A{50 + CONST}:AA{50 + CONST}","",{})
        
        ROW_DICT.update({key:0.18 for key in range(27 + i*OFFSET , 39 + i*OFFSET)})
        
        PAGEBREAKS.append(50 + CONST)

    
    worksheet.set_h_pagebreaks(PAGEBREAKS)
    annexUtils.set_row_dict(worksheet,ROW_DICT)
    
    # formatter.set_rows({0:2,1:2,2:2, 4:2})
    workbook.close()




if __name__ == "__main__":
    
    generate()
    #f1 = sys.argv[1]
    #f2 = sys.argv[2]
    
    #reader = rd.Reader (f1, "", f2)
    #matrix, labels, om, ol, heights = reader.getData()
    #model = mdl.Model(heights,matrix,labels, om, ol)
    
    #trans (model)

