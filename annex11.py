import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
import annexUtils
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
from openpyxl   import load_workbook


def generate (input_file = 'anexos/anteproyecto/anexo10.xlsx', output_file = "test11.xlsx") :
    
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet("Cotas PR")
    worksheetG = workbook.add_worksheet("Cotas G")
    worksheetT = workbook.add_worksheet("Cotas T")
    
    
    generatePR(workbook, worksheet, input_file)
    generateG(workbook, worksheetG, input_file)
    generateT(workbook, worksheetT, input_file)
    
    workbook.close()

##################################################################
#  __               _____ _____ _____ ___   _____  ____________  #
# /  |             /  __ \  _  |_   _/ _ \ /  ___| | ___ \ ___ \ #
# `| |    ______   | /  \/ | | | | |/ /_\ \\ `--.  | |_/ / |_/ / #
#  | |   |______|  | |   | | | | | ||  _  | `--. \ |  __/|    /  #
# _| |_            | \__/\ \_/ / | || | | |/\__/ / | |   | |\ \  #
# \___/             \____/\___/  \_/\_| |_/\____/  \_|   \_| \_| #
##################################################################

def generatePR (workbook,worksheet, input_file) :
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
    
    writer = Writer(workbook,worksheet)
    
    # FIXED CONTENT
    writer.merge(f"B2:D6","",Format.BORDER)
    writer.merge(
        f"E2:I5","COTAS DE PR",
        Format.BTOP,Format.BRIGHT,Format.SIZE(12),Format.BOLD,
        Format.CENTER, Format.VCENTER
    )
    
    writer.merge(
        f"E6:I6",
        "LÁMINA N° 2.903.3.I    FIGURA 2",
        Format.BBOTTOM,Format.BRIGHT,Format.SIZE(12),
        Format.BOLD,Format.CENTER
    )
    
    
    writer.write(f"J2","",Format.BLEFT)
    writer.merge(f"B7:I7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"J8:J12","",{"left":1})
    writer.merge(f"B13:I13", "",{"top":1})
    
    writer.write(f"B8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"G12:I12","FECHA:",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
  
    writer.merge(
        f'B14:C14',
        'PR',
        Format.BORDER,
        Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(
        f'D14:F14',
        'PROYECTO DEFINITIVO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(
        f'B15:B16',
        'DESDE',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'C15:C16',
        'HASTA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'D15:D16',
        'DESNIVEL\nIDA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'E15:E16',
        'DESNIVEL\nREGRESO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'F15:F16',
        'ERROR',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'G14:G16',
        'DESNIVEL\nPROMEDIO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'H14:H16',
        'COTA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'I14:I16',
        'ESTACION -\nPR',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(f'B17:I17','',Format.BORDER)
    
    wb = load_workbook(input_file)
    ws = wb.active
    
    max_row = ws.max_row
    first_row = True
    curr_row = 18
    DESDE = 1
    HASTA = 1
    BREAK = False
 
    POINT_A = ""
    POINT_B = ""
    DELTA_A = 0.0
    DELTA_B = 0.0
    ERR     = 0.0
    DELTA_MEAN = 0.0
    HEIGHT = 0.0
    
    
 
 
    
    while True:
        for col in ws.iter_cols(min_col=2,max_col=2,min_row=HASTA,max_row=max_row):
            for cell in col:
                if cell.value == "DESDE":
                    DESDE  = cell.row + 3
                    POINT_A = ws[f'B{DESDE}'].value
                    if first_row:
                        HEIGHT = ws[f'I{DESDE}'].value
                        writer.write(f'B{curr_row}','', Format.BORDER)
                        writer.write(f'C{curr_row}', POINT_A, Format.CENTER,Format.SIZE(10),Format.BORDER)
                        writer.write(f'D{curr_row}','', Format.BORDER)
                        writer.write(f'E{curr_row}','', Format.BORDER)
                        writer.write(f'F{curr_row}','', Format.BORDER)
                        writer.write(f'G{curr_row}','', Format.BORDER)
                        writer.write(f'H{curr_row}', HEIGHT, Format.CENTER,Format.SIZE(10),Format.BOLD,Format.BORDER)
                        writer.write(f'I{curr_row}', POINT_A, Format.CENTER,Format.SIZE(10),Format.BORDER)
                        curr_row += 1
                        first_row = False
                    BREAK = True
                    break
            if BREAK:
                break
        
        if not BREAK:
            break
        
        BREAK = False
        
        for col in ws.iter_cols(min_col=3,max_col=3,min_row=DESDE,max_row=max_row):
            for cell in col:
                if cell.value is not None:
                    HASTA = cell.row
                    POINT_B = ws[f'C{HASTA}'].value
                    BREAK = True
                    break
            if BREAK:
                break
            
        DELTA_A = ws[f'D{HASTA+1}'].value
        DELTA_B = ws[f'F{HASTA+1}'].value
        ERR     = ws[f'H{HASTA+1}'].value
        DELTA_MEAN = ws[f'D{HASTA+2}'].value
        HEIGHT = ws[f'I{HASTA}'].value
        
        writer.write(f'B{curr_row}', POINT_A,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'C{curr_row}', POINT_B,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'D{curr_row}', DELTA_A,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'E{curr_row}', DELTA_B,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'F{curr_row}', ERR,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'G{curr_row}', DELTA_MEAN,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'H{curr_row}', HEIGHT,Format.CENTER,Format.SIZE(10),Format.BOLD,Format.BORDER)
        writer.write(f'I{curr_row}', POINT_B,Format.CENTER,Format.SIZE(10),Format.BORDER)
        
        
        curr_row += 1
        BREAK = False
    
    
    writer.merge(f'B{curr_row}:I{curr_row}', "", Format.BTOP)
    
    formatter = Formatter(worksheet)
    col_width = {
        0:0.13,
        1:0.79,
        2:0.79,
        3:0.79,
        4:0.79,
        5:0.79,
        6:0.79,
        7:0.79,
        8:0.79,
        9:0.13,
    }
    # worksheet.autofit()
    formatter.set_cols(col_width)
    # formatter.set_rows({0:2,1:2,2:2, 4:2})
    
    
############################################################
#  __               _____ _____ _____ ___   _____   _____  #
# /  |             /  __ \  _  |_   _/ _ \ /  ___| |  __ \ #
# `| |    ______   | /  \/ | | | | |/ /_\ \\ `--.  | |  \/ #
#  | |   |______|  | |   | | | | | ||  _  | `--. \ | | __  #
# _| |_            | \__/\ \_/ / | || | | |/\__/ / | |_\ \ #
# \___/             \____/\___/  \_/\_| |_/\____/   \____/ #
############################################################
def generateG (workbook,worksheet,input_file) :
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
    
    writer = Writer(workbook,worksheet)

    # FIXED CONTENT
    writer.merge(f"B2:D6","",Format.BORDER)
    writer.merge(
        f"E2:I5","COTAS DE PR",
        Format.BTOP,Format.BRIGHT,Format.SIZE(12),Format.BOLD,
        Format.CENTER, Format.VCENTER
    )
    
    writer.merge(
        f"E6:I6",
        "LÁMINA N° 2.903.3.I    FIGURA 2",
        Format.BBOTTOM,Format.BRIGHT,Format.SIZE(12),
        Format.BOLD,Format.CENTER
    )
    
    
    writer.write(f"J2","",Format.BLEFT)
    writer.merge(f"B7:I7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"J8:J12","",{"left":1})
    writer.merge(f"B13:I13", "",{"top":1})
    
    writer.write(f"B8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"G12:I12","FECHA:",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
  
    writer.merge(
        f'B14:C14',
        'PR',
        Format.BORDER,
        Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(
        f'D14:F14',
        'PROYECTO DEFINITIVO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(
        f'B15:B16',
        'DESDE',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'C15:C16',
        'HASTA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'D15:D16',
        'DESNIVEL\nIDA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'E15:E16',
        'DESNIVEL\nREGRESO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'F15:F16',
        'ERROR',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'G14:G16',
        'DESNIVEL\nPROMEDIO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'H14:H16',
        'COTA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'I14:I16',
        'ESTACION -\nPR',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(f'B17:I17','',Format.BORDER)
    
    wb = load_workbook('anexos/anteproyecto/anexo10.xlsx')
    ws = wb["Nivelacion G y T"]
    max_row = ws.max_row
    first_row = True
    curr_row = 19
    DESDE = 1
    HASTA = 1
    BREAK = False
    CONTINUE = False
 
    POINT_A = ""
    POINT_B = ""
    DELTA_A = 0.0
    DELTA_B = 0.0
    ERR     = 0.0
    DELTA_MEAN = 0.0
    HEIGHT_A = 0.0
    HEIGHT_B = 0.0
    
    
 
 
    
    while True:
        for col in ws.iter_cols(min_col=2,max_col=2,min_row=HASTA,max_row=max_row):
            for cell in col:
                if cell.value == "DESDE":
                    DESDE  = cell.row + 3
                    POINT_A = ws[f'B{DESDE}'].value
                    BREAK = True
                    break
            if BREAK:
                break
        
        if not BREAK:
            break
        
        BREAK = False
        
        for col in ws.iter_cols(min_col=3,max_col=3,min_row=DESDE,max_row=max_row):
            for cell in col:
                if cell.value is not None and annexUtils.is_g(cell.value) :
                    HASTA = cell.row
                    POINT_B = ws[f'C{HASTA}'].value
                    BREAK = True
                    break
                if cell.value is not None and not annexUtils.is_g(cell.value) :
                    HASTA = cell.row
                    CONTINUE = True
            if BREAK:
                break
            
        if CONTINUE:
            CONTINUE = False
            continue
            
        DELTA_A = ws[f'D{HASTA+1}'].value
        DELTA_B = ws[f'F{HASTA+1}'].value
        ERR     = ws[f'H{HASTA+1}'].value
        DELTA_MEAN = ws[f'D{HASTA+2}'].value
        HEIGHT_A = ws[f'I{DESDE}'].value
        HEIGHT_B = ws[f'I{HASTA}'].value
        
        writer.write(f'B{curr_row-1}', '', Format.BORDER)
        writer.write(f'C{curr_row-1}', '', Format.BORDER)
        writer.write(f'D{curr_row-1}', '', Format.BORDER)
        writer.write(f'E{curr_row-1}', '', Format.BORDER)
        writer.write(f'F{curr_row-1}', '', Format.BORDER)
        writer.write(f'G{curr_row-1}', '', Format.BORDER)
        writer.write(f'H{curr_row-1}', HEIGHT_A,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'I{curr_row-1}', POINT_A,Format.CENTER,Format.SIZE(10),Format.BORDER)        
        
        writer.write(f'B{curr_row}', POINT_A,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'C{curr_row}', POINT_B,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'D{curr_row}', DELTA_A,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'E{curr_row}', DELTA_B,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'F{curr_row}', ERR,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'G{curr_row}', DELTA_MEAN,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'H{curr_row}', HEIGHT_B,Format.CENTER,Format.SIZE(10),Format.BOLD,Format.BORDER)
        writer.write(f'I{curr_row}', POINT_B,Format.CENTER,Format.SIZE(10),Format.BORDER)
        
        writer.write(f'B{curr_row+1}', '', Format.BORDER)
        writer.write(f'C{curr_row+1}', '', Format.BORDER)
        writer.write(f'D{curr_row+1}', '', Format.BORDER)
        writer.write(f'E{curr_row+1}', '', Format.BORDER)
        writer.write(f'F{curr_row+1}', '', Format.BORDER)
        writer.write(f'G{curr_row+1}', '', Format.BORDER)
        writer.write(f'H{curr_row+1}', '', Format.BORDER)
        writer.write(f'I{curr_row+1}', '', Format.BORDER) 
        
        curr_row += 3
        BREAK = False
    
    
    writer.merge(f'B{curr_row-1}:I{curr_row-1}', "", Format.BTOP)
    
    formatter = Formatter(worksheet)
    col_width = {
        0:0.13,
        1:0.79,
        2:0.79,
        3:0.79,
        4:0.79,
        5:0.79,
        6:0.79,
        7:0.79,
        8:0.79,
        9:0.13,
    }
    # worksheet.autofit()
    formatter.set_cols(col_width)
    # formatter.set_rows({0:2,1:2,2:2, 4:2})



############################################################
#  __               _____ _____ _____ ___   _____   _____  #
# /  |             /  __ \  _  |_   _/ _ \ /  ___| |_   _| #
# `| |    ______   | /  \/ | | | | |/ /_\ \\ `--.    | |   #
#  | |   |______|  | |   | | | | | ||  _  | `--. \   | |   #
# _| |_            | \__/\ \_/ / | || | | |/\__/ /   | |   #
# \___/             \____/\___/  \_/\_| |_/\____/    \_/   #
############################################################
def generateT(workbook,worksheet,input_file):
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
    
    writer = Writer(workbook,worksheet)

    # FIXED CONTENT
    writer.merge(f"B2:D6","",Format.BORDER)
    writer.merge(
        f"E2:I5","COTAS DE PR",
        Format.BTOP,Format.BRIGHT,Format.SIZE(12),Format.BOLD,
        Format.CENTER, Format.VCENTER
    )
    
    writer.merge(
        f"E6:I6",
        "LÁMINA N° 2.903.3.I    FIGURA 2",
        Format.BBOTTOM,Format.BRIGHT,Format.SIZE(12),
        Format.BOLD,Format.CENTER
    )
    
    
    writer.write(f"J2","",Format.BLEFT)
    writer.merge(f"B7:I7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"J8:J12","",{"left":1})
    writer.merge(f"B13:I13", "",{"top":1})
    
    writer.write(f"B8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"G12:I12","FECHA:",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
  
    writer.merge(
        f'B14:C14',
        'PR',
        Format.BORDER,
        Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(
        f'D14:F14',
        'PROYECTO DEFINITIVO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(
        f'B15:B16',
        'DESDE',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'C15:C16',
        'HASTA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'D15:D16',
        'DESNIVEL\nIDA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'E15:E16',
        'DESNIVEL\nREGRESO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'F15:F16',
        'ERROR',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'G14:G16',
        'DESNIVEL\nPROMEDIO',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'H14:H16',
        'COTA',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    writer.merge(
        f'I14:I16',
        'ESTACION -\nPR',
        Format.BORDER,Format.CENTER,Format.VCENTER,Format.SIZE(10)
    )
    
    writer.merge(f'B17:I17','',Format.BORDER)
    
    wb = load_workbook('anexos/anteproyecto/anexo10.xlsx')
    ws = wb["Nivelacion G y T"]
    max_row = ws.max_row
    first_row = True
    curr_row = 19
    DESDE = 1
    HASTA = 1
    BREAK = False
    CONTINUE = False
 
    POINT_A = ""
    POINT_B = ""
    DELTA_A = 0.0
    DELTA_B = 0.0
    ERR     = 0.0
    DELTA_MEAN = 0.0
    HEIGHT_A = 0.0
    HEIGHT_B = 0.0
    
    
 
 
    
    while True:
        for col in ws.iter_cols(min_col=2,max_col=2,min_row=HASTA,max_row=max_row):
            for cell in col:
                if cell.value == "DESDE":
                    DESDE  = cell.row + 3
                    POINT_A = ws[f'B{DESDE}'].value
                    BREAK = True
                    break
            if BREAK:
                break
        
        if not BREAK:
            break
        
        BREAK = False
        
        for col in ws.iter_cols(min_col=3,max_col=3,min_row=DESDE,max_row=max_row):
            for cell in col:
                if cell.value is not None and annexUtils.is_t(cell.value) :
                    HASTA = cell.row
                    POINT_B = ws[f'C{HASTA}'].value
                    BREAK = True
                    break
                if cell.value is not None and not annexUtils.is_t(cell.value) :
                    HASTA = cell.row
                    BREAK    = True
                    CONTINUE = True
                    break
            if BREAK:
                break
            
        if CONTINUE:
            CONTINUE = False
            BREAK = False
            continue
            
        DELTA_A = ws[f'D{HASTA+1}'].value
        DELTA_B = ws[f'F{HASTA+1}'].value
        ERR     = ws[f'H{HASTA+1}'].value
        DELTA_MEAN = ws[f'D{HASTA+2}'].value
        HEIGHT_A = ws[f'I{DESDE}'].value
        HEIGHT_B = ws[f'I{HASTA}'].value
        
        writer.write(f'B{curr_row-1}', '', Format.BORDER)
        writer.write(f'C{curr_row-1}', '', Format.BORDER)
        writer.write(f'D{curr_row-1}', '', Format.BORDER)
        writer.write(f'E{curr_row-1}', '', Format.BORDER)
        writer.write(f'F{curr_row-1}', '', Format.BORDER)
        writer.write(f'G{curr_row-1}', '', Format.BORDER)
        writer.write(f'H{curr_row-1}', HEIGHT_A,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'I{curr_row-1}', POINT_A,Format.CENTER,Format.SIZE(10),Format.BORDER)        
        
        writer.write(f'B{curr_row}', POINT_A,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'C{curr_row}', POINT_B,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'D{curr_row}', DELTA_A,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'E{curr_row}', DELTA_B,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'F{curr_row}', ERR,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'G{curr_row}', DELTA_MEAN,Format.CENTER,Format.SIZE(10),Format.BORDER)
        writer.write(f'H{curr_row}', HEIGHT_B,Format.CENTER,Format.SIZE(10),Format.BOLD,Format.BORDER)
        writer.write(f'I{curr_row}', POINT_B,Format.CENTER,Format.SIZE(10),Format.BORDER)
        
        writer.write(f'B{curr_row+1}', '', Format.BORDER)
        writer.write(f'C{curr_row+1}', '', Format.BORDER)
        writer.write(f'D{curr_row+1}', '', Format.BORDER)
        writer.write(f'E{curr_row+1}', '', Format.BORDER)
        writer.write(f'F{curr_row+1}', '', Format.BORDER)
        writer.write(f'G{curr_row+1}', '', Format.BORDER)
        writer.write(f'H{curr_row+1}', '', Format.BORDER)
        writer.write(f'I{curr_row+1}', '', Format.BORDER) 
        
        curr_row += 3
        BREAK = False
    
    
    writer.merge(f'B{curr_row-1}:I{curr_row-1}', "", Format.BTOP)
    
    formatter = Formatter(worksheet)
    col_width = {
        0:0.13,
        1:0.79,
        2:0.79,
        3:0.79,
        4:0.79,
        5:0.79,
        6:0.79,
        7:0.79,
        8:0.79,
        9:0.13,
    }
    # worksheet.autofit()
    formatter.set_cols(col_width)
    # formatter.set_rows({0:2,1:2,2:2, 4:2})









if __name__ == "__main__":
    generate()
    
    
