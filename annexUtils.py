import re
from datetime import datetime
from openpyxl.utils import column_index_from_string, get_column_letter

INCH_COL = 10
INCH_ROW = 72
OFFSET   = 38
PAGEBREAKS = []

def letter_of_int (i) :
    return get_column_letter(i)

def int_of_letter (l) :
    return column_index_from_string(l)

def coordinate_of_ints (i,j) :
    return f'{letter_of_int(i)}:{j}'

class Writer :
    def __init__ (self,workbook,worksheet):
        self.workbook = workbook
        self.worksheet = worksheet
    
    def set_worksheet (self,ws):
        self.worksheet = ws
    
    def merge (self,ran,dat,*dic):
        
        
        CELL_FORMAT = {}
        for f in dic :
            CELL_FORMAT = CELL_FORMAT | f
            
        self.worksheet.merge_range(
            ran,
            dat,
            self.workbook.add_format(CELL_FORMAT)
        )
    
    def write (self,cell,dat,*dic):
        
        CELL_FORMAT = {}
        for f in dic :
            CELL_FORMAT = CELL_FORMAT | f
            
        self.worksheet.write(
            cell,
            dat,
            self.workbook.add_format(CELL_FORMAT)
        )


class Format :
    
    # Functions
    def SIZE(n):
        return {"font_size":n}
    
    # Variables
    BOLD   = {"bold":True}
    ITALIC = {"italic":True}
    DEC    = {'num_format': '#,##0.0000000000'}
    NUM    = {'num_format': '#,##0.000'}
    TOP    = {"top" :1}
    BTOP   = {"top" :1}
    BBOTTOM = {"bottom" :1}
    BOTTOM = {"bottom":1}
    BLEFT = {"left": 1}
    BRIGHT = {"right": 1}
    BORDER = {"border":1}
    BORDER_THICK = {"border":2}
    LEFT = {"align":"left"}
    RIGHT = {"align":"right"}
    CENTER = {"align":"center"}
    VCENTER = {"valign":"vcenter"}




class Formatter :
    
    def __init__(self, ws):
        self.ws = ws
        
    
    def set_cols(self,dic):
        for col in dic:
            size = dic[col]
            self.ws.set_column(col,col,size*INCH_COL) 
            
    
    def set_rows(self,dic):
        for row in dic:
            size = dic[row]
            self.ws.set_row(row,size*INCH_ROW)

def set_column (ws,widths,INCH_COL=10):
    for col , w in enumerate(widths):
        ws.set_column(col,col,w*INCH_COL)

def set_row (ws, heights) :
    for row , h in enumerate(heights) :
        ws.set_row(row,h*INCH_ROW)

def set_row_dict (ws, heights={}) :
    for h in heights:
        ws.set_row(h, heights[h] * INCH_ROW)



def is_g (s):
    if s.startswith("G"):
        return True
    else:
        False


def is_t (s):
    if s.startswith("T"):
        return True
    else:
        return False

def curr_date (opt=0) :
    
    if opt  == 0:
        MES = {
            1 : "ENERO",
            2 : "FEBRERO",
            3 : "MARZO",
            4 : "ABRIL",
            5 : "MAYO",
            6 : "JUNIO",
            7 : "JULIO",
            8 : "AGOSTO",
            9 : "SEPTIEMBRE",
            10 : "OCTUBRE",
            11 : "NOVIEMBRE"
        }
        return f'{MES[datetime.now().month]} {datetime.now().year}'
    
    else :
        MES  = {
            1 : "ENE",
            2 : "FEB",
            3 : "MAR",
            4 : "ABR",
            5 : "MAY",
            6 : "JUN",
            7 : "JUL",
            8 : "AGO",
            9 : "SEP",
            10 : "OCT",
            11 : "NOV"
        }
        return f'{MES[datetime.now().month]} {datetime.now().year}'

class Scanner :
    
    def __init__ (self,ws) :
        
        self.ws = ws
        self.row = self.__find_row__()
        
        self.DATUM = None
        self.ZONA  = None
        self.MC    = None
        
        self.PTL_COLS = []
        
        self.ALTURA_PTL        = []
        self.MERIDIANO_CENTRAL = []
        self.NORTE_FALSO       = []
        self.ESTE_FALSO        = []
        self.FACTOR_ESCALA     = []
        
        self.PRO       = None
        self.EST       = None
        self.GEO_S     = None
        self.GEO_W     = None
        self.UTM_N     = None
        self.UTM_E     = None
        self.GEO_X     = None
        self.GEO_Y     = None
        self.GEO_Z     = None
        self.PTL_N     = []
        self.PTL_E     = []
        self.ELIP      = None
        self.COTA_ORTO = None
        self.COTA_GEO  = None
        
        self.__find_field__()
        self.__init_ptl__ ()
        self.__init_global__()
    
    def get_pro(self,row):
        return self.ws.cell(column=self.PRO,row=row).value
    
    def get_est(self,row):
        return self.ws.cell(column=self.EST,row=row).value
    
    def get_geo_s(self,row):
        return self.ws.cell(column=self.GEO_S,row=row).value
    
    def get_geo_w (self,row):
        return self.ws.cell(column=self.GEO_W,row=row).value
    
    def get_utm_n (self,row):
        return self.ws.cell(column=self.UTM_N,row=row).value
    
    def get_utm_e  (self,row):
        return self.ws.cell(column=self.UTM_E,row=row).value
    
    def get_geo_x  (self,row):
        return self.ws.cell(column=self.GEO_X,row=row).value
    
    def get_geo_y  (self,row):
        return self.ws.cell(column=self.GEO_Y,row=row).value
    
    def get_geo_z  (self,row):
        return self.ws.cell(column=self.GEO_Z,row=row).value
    
    def get_elip  (self,row):
        return self.ws.cell(column=self.ELIP,row=row).value
    
    def get_cota_orto  (self,row):
        return self.ws.cell(column=self.COTA_ORTO,row=row).value
    
    def get_cota_geo  (self,row):
        return self.ws.cell(column=self.COTA_GEO,row=row).value
    
 
    def __init_global__ (self) :
        for col in self.ws.iter_cols(min_row = 0, max_row = self.row):
            for cell in col:
                r = cell.row
                c = cell.column
                if bool(re.match(r"^datum", str(cell.value), re.I)):
                    self.DATUM = self.ws.cell(row=r,column=c+1).value
                    continue
                if bool(re.match(r"^zona", str(cell.value), re.I)):
                    self.ZONA = self.ws.cell(row=r,column=c+1).value
                    continue
                if bool(re.match(r"^mc", str(cell.value), re.I)):
                    self.MC = self.ws.cell(row=r, column=c+1).value
                    break
    
    
    def get_t_rows (self) :
        t_rows = []
        for col in self.ws.iter_cols(min_col=self.EST, max_col=self.EST, min_row=self.row):
                for cell in col:
                    if bool(re.match(r"T-", str(cell.value), re.I)):
                        t_rows.append(cell.row)
        return t_rows
    
    def get_g_rows (self) :
        g_rows = []
        for col in self.ws.iter_cols(min_col=self.EST, max_col=self.EST, min_row=self.row):
                for cell in col:
                    if bool(re.match(r"G-", str(cell.value), re.I)):
                        g_rows.append(cell.row)
        return g_rows
        
    
    def get_poligonal_num (self, row):
        target_cell = self.ws.cell(row = row, column = self.PRO)
        for merged_range in self.ws.merged_cells.ranges:
            if target_cell.coordinate in merged_range:
                cell = self.ws.cell(row=merged_range.min_row, column=merged_range.min_col)
                match = re.search(r'\d+', str(cell.value))
                if match:
                    return int(match.group())
                else:
                    return -1
    
    def __init_ptl__ (self):
        for i in self.PTL_COLS:
            for col in self.ws.iter_cols(min_col=i,max_col=i,min_row=1,max_row=self.row):
                for cell in col:
                    if bool(re.match(r"altura", str(cell.value), re.I)):
                        self.ALTURA_PTL.append(self.ws.cell(column = cell.column + 1 , row = cell.row).value)
                        continue
                    if bool(re.match(r"meridiano", str(cell.value), re.I)):
                        self.MERIDIANO_CENTRAL.append(self.ws.cell(column = cell.column + 1 , row = cell.row).value)
                        continue
                    if bool(re.match(r"norte", str(cell.value), re.I)):
                        self.NORTE_FALSO.append(self.ws.cell(column = cell.column + 1 , row = cell.row).value)
                        continue
                    if bool(re.match(r"este", str(cell.value), re.I)):
                        self.ESTE_FALSO.append(self.ws.cell(column = cell.column + 1 , row = cell.row).value)
                        continue
                    if bool(re.match(r"factor", str(cell.value), re.I)):
                        self.FACTOR_ESCALA.append(self.ws.cell(column = cell.column + 1 , row = cell.row).value)
                        continue
    
 
    def __find_row__ (self) :
        for col in self.ws.iter_cols(min_col=1):
            for cell in col:
                if cell.value == "MASTER":
                    return cell.row
        raise Exception("No row starting with \'MASTER\' cell")
    
    def __find_field__ (self) :
        for row in self.ws.iter_rows(min_row=self.row,max_row=self.row):
            for cell in row:
                if cell.value == "PRO":
                    self.PRO = cell.column
                    continue
                if cell.value == "EST":
                    self.EST = cell.column
                    continue
                if cell.value == "GEO-S":
                    self.GEO_S = cell.column
                    continue
                if cell.value == "GEO-W":
                    self.GEO_W = cell.column
                    continue
                if cell.value == "UTM-N":
                    self.UTM_N = cell.column
                    continue
                if cell.value == "UTM-E":
                    self.UTM_E = cell.column
                    continue
                if cell.value == "GEO-X":
                    self.GEO_X = cell.column
                    continue
                if cell.value == "GEO-Y":
                    self.GEO_Y = cell.column
                    continue
                if cell.value == "GEO-Z":
                    self.GEO_Z = cell.column
                    continue
                if  bool(re.match(r"^PTL\d*-N", str(cell.value))):
                    self.PTL_COLS.append(cell.column)
                    self.PTL_N.append(cell.column)
                    continue
                if  bool(re.match(r"^PTL\d*-E", str(cell.value))):
                    self.PTL_E.append(cell.column)
                    continue
                if cell.value == "ELIP":
                    self.ELIP = cell.column
                    continue
                if cell.value == "COTA-ORTO":
                    self.COTA_ORTO = cell.column
                    continue
                if cell.value == "COTA-GEO":
                    self.COTA_GEO =  cell.column
                    break


if __name__ == "__main__":
    
    print(curr_date())
