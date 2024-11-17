

import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
import annexUtils
from openpyxl import load_workbook
import os
import glob
import annexImg

OFFSET   = 30
PAGEBREAKS = []

def generate (input_file='anexos/anteproyecto/anexo1.xlsx',output_file="test5.xlsx", src_dir="img", src_dir2="img_geo") :
    
    print(f'\n\nGeneración de {output_file} en curso ...')
    
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet("Fichas PR")
    writer = Writer(workbook,worksheet)
    wb = load_workbook(input_file)
    ws = wb.active
    scanner = annexUtils.Scanner(ws)
    dst_dir = None
    dst_dir2 = None
    
    COL_WIDTHS = [
        0.10, #A 
        0.30, #B
        0.70, #C
        0.25, #D
        0.39, #E
        0.18, #F
        0.19, #G
        0.19, #H
        0.25, #I
        0.22, #J
        0.30, #K
        0.25, #L
        0.15, #M
        0.42, #N
        0.15, #O
        0.20, #P
        0.20, #Q
        0.33, #R
        0.25, #S
        0.15, #T
        0.25, #U
        0.15, #V
        0.33, #W
        0.28, #X
        0.46, #Y
        0.25, #Z
        0.10, #AA
    ]
    
    ROW_DICT = {
        0 :0.1,  
        6 :0.1,  
        7 :0.18, 
        8 :0.18, 
        9 :0.19, 
        10:0.1,
        12:0.12,
        13:0.12
        
    }
    
    #worksheet.autofit()
    annexUtils.set_column(worksheet,COL_WIDTHS)
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    
    # FIXED CONTENT
    writer.merge(f"B2:F6","",Format.BORDER)
    writer.merge(f"G2:Z3","MONOGRAFÍAS DE PR",
                 {"top":1, "right":1,"font_size":12,"bold":True,"align":"center","valign":"vcenter"})
    
    writer.merge(f"G4:Z6","FORMULARIO N° 2.903.3.I   FIGURA 3",
                 {"bottom":1,"right":1, "font_size":12,"bold":True,"align":"center","valign":"vcenter"})
    
    writer.write(f"AA2","",Format.BLEFT)
    
    writer.merge(f"B7:Z7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"AA8:AA12","",{"left":1})
    writer.merge(f"B13:Z13", "",{"top":1})
    
    writer.merge(f"B8:C8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B9:C9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B10:C10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B12:C12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"U12:Z12",f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
    
    writer.merge(f'D8:Y8', scanner.PROYECTO, Format.LEFT,Format.SIZE(10))
    writer.merge(f'D9:Y9',scanner.SECTOR,Format.SIZE(10))
    writer.merge(f'D10:Y10', scanner.TRAMO,Format.SIZE(10))
    writer.merge(f'D12:R12',scanner.REALIZADO,Format.SIZE(10))
    
    
  
    t_rows = scanner.get_t_rows()
    i = 0
    
    if src_dir:
        dst_dir = annexImg.annex5_process(src_dir)
        
    if src_dir2:
        dst_dir2 = annexImg.annex5_process_geo(src_dir2)
    
    
    # LOOP THE FOLLOWING CELLS 
    for r in t_rows :
        
        #CELL_nombre = ws[f'C{r}'].value # DONE
        POINT = scanner.get_est(r)
        
        CELL_f      = scanner.get_geo_s(r)
        CELL_l      = scanner.get_geo_w(r)
        CELL_h      = scanner.get_elip(r)
        
        POLY_NUM    = scanner.get_poligonal_num(r)
        
        CELL_X      = scanner.get_geo_x(r)
        CELL_Y      = scanner.get_geo_y(r)
        CELL_Z      = scanner.get_geo_z(r)
        
        CELL_N      = scanner.get_utm_n(r)
        CELL_E      = scanner.get_utm_e(r)
        
        CELL_altura = scanner.get_cota_orto(r)
        CELL_cota   = scanner.get_cota_geo(r)
        
        CELL_dm     = scanner.get_dm(r)
        CELL_dist   = scanner.get_dist(r)
        
        
        # CELL_NL     = ws[f'K{r}'].value
        # CELL_EL     = ws[f'L{r}'].value
        # CELL_MCL    = ws['H9'].value
        # CELL_Ko     = ws['H12'].value
        
        
        
        writer.write(f"B{15 + i * OFFSET}","Identificación del Punto",Format.BOLD, Format.ITALIC, Format.SIZE(10))
        writer.write(f"F{15 + i * OFFSET}", "PR:", Format.SIZE(10))
        writer.write(f"K{15 + i * OFFSET}", "Dm. Ref.:",Format.SIZE(10))
        writer.write(f"R{15 + i * OFFSET}", "Lado:")
        
        writer.write(
            f'F{16 + i * OFFSET}',
            "Cota:",
            Format.SIZE(10)
        )
        
        writer.merge(f"H{16 + i * OFFSET}:I{16 + i * OFFSET}", "23.43", Format.NUM)
        writer.write(f"J{16 + i * OFFSET}", "m")
        
        writer.write(
            f'L{16 + i * OFFSET}',
            "Coordenadas de Navegación UTM:",
            Format.SIZE(10)
        )
        
        writer.write(
            f'U{16 + i * OFFSET}',
            "N:"
        )
        
        writer.write(
            f'U{17 + i * OFFSET}',
            "E:"
        )
        
        
        
        writer.merge(
            f"H{15 + i * OFFSET}:I{15 + i * OFFSET}",
            POINT,
            Format.SIZE(11), Format.BOTTOM, Format.CENTER
        )
        
        
        writer.merge(f"M{15 + i * OFFSET}:P{15 + i * OFFSET}", CELL_dm,Format.CENTER,Format.SIZE(10),Format.BOTTOM,Format.NUM2)
        writer.write(f"U{15 + i * OFFSET}", "D", Format.CENTER, Format.BBOTTOM)
        writer.write(f"W{15 + i * OFFSET}","FECHA:",Format.SIZE(10))
        writer.merge(f"Y{15 + i * OFFSET}:Z{15 + i * OFFSET}", annexUtils.curr_date(1),Format.BOTTOM,Format.CENTER,Format.SIZE(10))
        
        
        
        
        
        if src_dir:
            cleaned_point = POINT.replace("-", "")
            
            img_path_a = os.path.join(dst_dir, f'{POINT}*', f'{cleaned_point}_a.*')
            img_path_p = os.path.join(dst_dir, f'{POINT}*', f'{cleaned_point}_p.*')
            
            match_a = glob.glob(img_path_a)
            match_p = glob.glob(img_path_p)
            
            if match_a:
                worksheet.insert_image(f'B{33 + i * OFFSET}', match_a[0] , {'object_position': 1})
            else :
                writer.merge(
                    f"B{33 + i * OFFSET}:F{40 + i * OFFSET}","Fotografía\nDetalle",
                    Format.BORDER,
                    Format.CENTER,Format.VCENTER)
            
            if match_p:
                worksheet.insert_image(f'B{19 + i * OFFSET}', match_p[0],  {'object_position': 1})
            else:
                writer.merge(
                    f"B{19 + i * OFFSET}:L{31 + i * OFFSET}","Fotografía\nPanorámica",
                    Format.BORDER,
                    Format.CENTER,Format.VCENTER)
        
     
        if src_dir2 :
            cleaned_point = POINT.replace("-", "")
            img_path_g = os.path.join(dst_dir2, f'{cleaned_point}_g.*')
            match_g = glob.glob(img_path_g)
            
            if match_g:
                worksheet.insert_image(f'N{19 + i * OFFSET}', match_g[0] , {'object_position': 1})
            else:
                writer.merge(
                    f"N{19 + i * OFFSET}:Z{31 + i * OFFSET}","Vista\nAérea",
                    Format.BORDER,
                    Format.CENTER,Format.VCENTER)
        
        
        writer.merge(f"H{41-8 + i * OFFSET}:Z{41-8 + i * OFFSET}","Descripción",Format.BOTTOM,Format.LEFT,Format.SIZE(10))
        
        writer.write(f'I{43-8 + i * OFFSET}', "Materialidad:",Format.SIZE(9))
        writer.write(f'I{44-8 + i * OFFSET}', "Dimensiones:",Format.SIZE(9))
        writer.merge(f'I{45-8 + i * OFFSET}:N{45-8 + i * OFFSET}', "Distancia a la Ruta:",Format.SIZE(9))
        writer.merge(f'O{45-8 + i * OFFSET}:Q{45-8 + i * OFFSET}', CELL_dist, Format.SIZE(9), Format.NUM2)
        writer.write(f'R{45-8 + i * OFFSET}', "m.", Format.SIZE(9), Format.LEFT)
        
        writer.write(f'L{43-8 + i * OFFSET}', "MONOLITO DE HORMIGÓN, PINTADO DE AMARILLO", Format.SIZE(9))
        writer.write(f'L{44-8 + i * OFFSET}', "D: 15 cm. FIERRO ESTRIADO 12 mm.", Format.SIZE(9))
        
        writer.merge(f"G{42-8 + i * OFFSET}:G{48-8 + i * OFFSET}","",Format.BRIGHT)
        writer.merge(f"AA{42-8 + i * OFFSET}:AA{48-8 + i * OFFSET}","",Format.BLEFT)
        writer.merge(f"H{49-8 + i * OFFSET}:Z{49-8 + i * OFFSET}","",Format.TOP)
        writer.merge(f"A{50-8 + i * OFFSET}:AA{50-8 + i * OFFSET}","",{})
        
        # ROW_DICT.update({15 + i * OFFSET : 0.25})
        # ROW_DICT.update({16 + i * OFFSET : 0.45})
        # ROW_DICT.update({17 + i * OFFSET : 0.14})
        # ROW_DICT.update({18 + i * OFFSET : 0.14})
        
        ROW_DICT.update({key:0.16 for key in range(19 + i*OFFSET , 31 + i*OFFSET)})  # - 8
        ROW_DICT.update({key:0.175 for key in range(33 + i*OFFSET , 40 + i*OFFSET)})  
        
        PAGEBREAKS.append(50-8 + i * OFFSET)
        
        i += 1
    
    worksheet.set_h_pagebreaks(PAGEBREAKS)
    annexUtils.set_row_dict(worksheet,ROW_DICT)
    
    # formatter.set_rows({0:2,1:2,2:2, 4:2})
    workbook.close()
    
    print(f'\nGeneración de {output_file} completada\n')




if __name__ == "__main__":
    generate(input_file='/home/jstvns/axis/axis/anexos/anteproyecto/anexo1.xlsx',output_file="../test9.xlsx", src_dir = '', src_dir2= '')
