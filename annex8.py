import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
import annexUtils
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
from openpyxl   import load_workbook

class Polygone :
    def __init__ (self, coordinate, min_row, max_row, value):
        self.coordinate = coordinate
        self.min_row = min_row
        self.max_row = max_row
        self.value = value
        
    def __lt__ (self,p) :
        return self.min_row < p.min_row
    
    def __le__ (self,p) :
        return self.min_row <= p.min_row
    
    def __str__ (self)  :
        return f'coor: {self.coordinate}\nmin_row: {self.min_row}\nmax_row: {self.max_row}\nval: {self.value}'


def generate (input_file = 'anexos/anteproyecto/anexo1.xlsx', output_file='test8.xlsx') :
    
    workbook = xlsxwriter.Workbook(output_file)
    wb = load_workbook(input_file)
    ws = wb.active
    
    column = 'B'
    
    polygones = [] 
    
    for merged_range in ws.merged_cells.ranges:
        
        if merged_range.min_col == openpyxl.utils.column_index_from_string(column):
            
            top_left_cell = ws.cell(row=merged_range.min_row, column=merged_range.min_col)
            if top_left_cell.value.startswith("POLIGONAL"):
                polygone = Polygone(
                    top_left_cell.coordinate,
                    merged_range.min_row,
                    merged_range.max_row,
                    top_left_cell.value
                )
                polygones.append(polygone)
    
    polygones.sort()
    
    # for p in polygones:
    #     print("")
    #     print(p)
    #     print("")
    
    
    COL_WIDTHS = [
        0.13, #A
        0.67, #B
        0.13, #C
        0.20, #D
        0.15, #E
        0.55, #F
        0.20, #G
        0.55, #H
        0.14, #I
        0.30, #J
        0.14, #K
        0.25, #L
        0.30, #M
        0.20, #N
        0.30, #O
        0.20, #P
        0.30, #Q
        0.20, #R
        0.20, #S
        0.20, #T
        0.20, #U
        0.70, #V
        0.13, #W
    ]
    
    ROW_HEIGHTS = [
        0.17, #1
        0.17, #2
        0.23, #3
        0.17, #4
        0.07, #5
        0.19, #6
        0.07, #7
        0.17, 0.17, 0.17, # 8,9,10
        0.07, #11
    ]
    
    for p in polygones:
        
        worksheet = workbook.add_worksheet(p.value)
        
        annexUtils.set_column(worksheet,COL_WIDTHS)
        annexUtils.set_row(worksheet,ROW_HEIGHTS)
        
        worksheet.hide_gridlines(2)
        worksheet.set_portrait()
        worksheet.set_page_view(2)
        worksheet.set_paper(9)
        worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
        
        writer = Writer(workbook,worksheet)
        
        # FIXED CONTENT
        writer.merge(f"B2:G6","",Format.BORDER)
        writer.merge(
            f"H2:V5","COORDENADAS DE VÉRTICES DEL STC",
            Format.BTOP,Format.BRIGHT,Format.SIZE(12),Format.BOLD,
            Format.CENTER, Format.VCENTER
        )
        
        writer.merge(
            f"H6:V6",
            "FORMULARIO N° 2.303.104.B",
            Format.BBOTTOM,Format.BRIGHT,Format.SIZE(12),
            Format.BOLD,Format.CENTER
        )
        
        
        writer.write("W2","",Format.BLEFT)
        writer.merge("B7:V7","",{"bottom":1})
        writer.merge("A8:A12","",{"right":1})
        writer.merge("W8:W12","",{"left":1})
        writer.merge("B13:V13", "",{"top":1})
        
        writer.write("B8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
        writer.write("B9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
        writer.write("B10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
        writer.write("B12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
        writer.merge("S12:V12","FECHA:",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
        
        writer.merge(
            'B14:Q14',
            'POLIGONAL Nº: 6      Tipo: Principal',
            Format.SIZE(9),
            Format.BOLD,Format.BORDER,Format.CENTER,Format.VCENTER,
        )
        
        writer.merge(
            'B15:B18','Vértice',Format.SIZE(9),
            Format.BOLD,Format.CENTER,Format.VCENTER, Format.BORDER
        )
        
        writer.merge( 'D15:I15','PTL',Format.SIZE(9),Format.BOLD,Format.CENTER)
        writer.merge( 'D16:E16','MCL:',Format.SIZE(9), Format.BOLD, Format.LEFT)
        writer.merge( 'F16:I16','72.23',Format.SIZE(9), Format.BOLD, Format.LEFT)
        writer.merge( 'D17:E17','Kptl:',Format.SIZE(9), Format.BOLD, Format.LEFT)
        writer.merge( 'F17:I17','61923784',Format.SIZE(9), Format.BOLD, Format.LEFT)
        writer.merge( 'J15:J17','', Format.BRIGHT)
        writer.merge( 'C18:J18','', Format.BBOTTOM,Format.BRIGHT)
        
        writer.merge( 'L16:P16','UTM',Format.SIZE(9), Format.BOLD, Format.CENTER)
        writer.merge( 'L17:P17','Huso:18?',Format.SIZE(9), Format.BOLD, Format.CENTER)
        writer.merge( 'K18:Q18', '', Format.BBOTTOM,Format.BRIGHT)
        writer.merge( 'Q15:Q17', '', Format.BRIGHT)
        
        curr_row = 19
        offset   = 4
        
        for i in range(p.min_row+1, p.max_row):
            
            POINT = ws[f'C{i}'].value
            NL    = ws[f'K{i}'].value
            EL    = ws[f'L{i}'].value
            N     = ws[f'F{i}'].value
            E     = ws[f'G{i}'].value
            COTA  = ws[f'O{i}'].value
            H     = ws[f'N{i}'].value
            
            writer.merge(
                f'B{curr_row}:B{curr_row+3}',POINT,Format.SIZE(9),
                Format.BOLD,Format.CENTER,Format.VCENTER, Format.BORDER
            )
            
            writer.merge(f'D{curr_row}:E{curr_row}', "NL:",Format.SIZE(10))
            writer.merge(f'D{curr_row+1}:E{curr_row+1}', "EL:",Format.SIZE(10))
            writer.merge(f'D{curr_row+2}:F{curr_row+2}', "Cota (nivelada):",Format.SIZE(9))
            
            writer.merge(f'F{curr_row}:H{curr_row}',NL,Format.RIGHT,Format.SIZE(10))
            writer.merge(f'F{curr_row+1}:H{curr_row+1}',EL,Format.RIGHT,Format.SIZE(10))
            writer.merge(f'G{curr_row+2}:H{curr_row+2}',COTA,Format.RIGHT,Format.SIZE(10))
            
            writer.write(f'I{curr_row}','m')
            writer.write(f'I{curr_row+1}','m')
            writer.write(f'I{curr_row+2}','m')
            
            writer.merge(f'C{curr_row+3}:J{curr_row+3}','',Format.BOTTOM,Format.BRIGHT)
            writer.merge(f'J{curr_row}:J{curr_row+2}','',Format.BRIGHT)
            
            
            
            writer.write(f'L{curr_row}',"N:",Format.LEFT,Format.SIZE(10))
            writer.write(f'L{curr_row+1}',"E:",Format.LEFT,Format.SIZE(10))
            writer.merge(f'L{curr_row+2}:M{curr_row+2}',"H(model):",Format.LEFT,Format.SIZE(9))
            
            writer.merge(f'M{curr_row}:O{curr_row}', N, Format.RIGHT,Format.SIZE(10))
            writer.merge(f'M{curr_row+1}:O{curr_row+1}', E, Format.RIGHT,Format.SIZE(10))
            writer.merge(f'N{curr_row+2}:O{curr_row+2}', H, Format.RIGHT,Format.SIZE(10))
            
            writer.merge(f'Q{curr_row}:Q{curr_row+2}', '', Format.BRIGHT)
            writer.merge(f'K{curr_row+3}:Q{curr_row+3}', '',Format.BBOTTOM,Format.BRIGHT)
            
            
            writer.write(f'P{curr_row}'  , 'm', Format.RIGHT)
            writer.write(f'P{curr_row+1}', 'm', Format.RIGHT)
            writer.write(f'P{curr_row+2}', 'm', Format.RIGHT)
            
            
            curr_row += offset
            
        
        writer.merge(f'B{curr_row}:Q{curr_row}','',Format.BTOP)
        
        
        
        
    
    workbook.close()

if __name__ == "__main__":
    generate()
