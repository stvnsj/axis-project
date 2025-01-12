import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
import annexUtils
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
from openpyxl   import load_workbook
import level

ROW_DICT = {}

def generate (
        input1 = "",
        input2 = "",
        output_file = "/home/jstvns/axis/Anexos Formulas/groundline.xlsx") :
    
    workbook   = xlsxwriter.Workbook(output_file)
    worksheet  = workbook.add_worksheet("SHEET1")
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
    
    writer = Writer(workbook,worksheet)
    
    writer.cell (2,8,"PROYECTO:",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.cell (2,9,"SECTOR:", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.cell (2,10,"TRAMO:", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.cell (2,11,"REALIZADO:",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.range(35,45,12,12,f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
    
    writer.cell(10,8,"NOMBRE DEL PROYECTO")
    writer.cell(10,9,"SECTOR DEL PROYECTO")
    writer.cell(10,10,"TRAMO DEL PROYECTO")
    writer.cell(10,11,"REALIZADO POR EQC")

    # Project data
    writer.range(2,6,14,14,"Tramo N°:"           ,Format.SIZE(9))
    writer.range(7,11,14,14,"34"          ,Format.SIZE(9),Format.CENTER) #PROGRAM
    writer.range(12,16,14,14,"Dm. Inicio:"        ,Format.SIZE(9))
    writer.range(17,23,14,14,"322,520.00"    ,Format.SIZE(9),Format.CENTER) #PROGRAM
    writer.range(24,27,14,14,"Dm. Fin:"           ,Format.SIZE(9))
    writer.range(28,33,14,14,"123,321.23"  ,Format.SIZE(9),Format.CENTER) #PROGRAM
    writer.range(34,42,14,14,"Longitud del Tramo:",Format.SIZE(9))
    writer.range(43,49,14,14,"123,123.32"  ,Format.SIZE(9),Format.CENTER) #PROGRAM
    
    index = 20
    
    writer.range(2,17,index,index, "Línea de Tierra Transversal", Format.BOLD, Format.SIZE(10))
    
    index+=1
    writer.range(2,14,index,index, "N° puntos contrastados:")
    writer.range(15,18,index,index, "344") #PROGRAM
    writer.range(19, 30, index, index, "Puntos en Tolerancia:")
    writer.range(31, 35, index, index, "239") #PROGRAM
    writer.range(36,37, index, index, "%")
    writer.range(38,42, index, index, "89.11") #PROGRAM
    
    index+=5
    writer.range(2,5,index,index,  "Perfil N°",Format.SIZE(9))
    writer.range(6,9,index,index,  "Dm.",Format.SIZE(9))
    writer.range(10,13,index,index,"Lado",Format.SIZE(9))
    
    writer.range(14,18,index-1,index-1,"Dist.",Format.SIZE(9))
    writer.range(14,18,index,index,"al Eje",Format.SIZE(9))
    
    writer.range(19,23,index-1,index-1,"Cota",Format.SIZE(9))
    writer.range(19,23,index,index, "Control",Format.SIZE(9))
    
    writer.range(24,28,index-1,index-1,"Cota",Format.SIZE(9))
    writer.range(24,28,index,index,"Estudio",Format.SIZE(9))
    
    writer.range(29,33,index-1,index-1,"Tipo",Format.SIZE(9))
    writer.range(29,33,index,index,"Sup. (1-4)",Format.SIZE(9))
    
    writer.range(34,38,index,index, "Tol. (m)",Format.SIZE(9))
    writer.range(39,42,index,index, "Dif. (m)",Format.SIZE(9))
    writer.range(43,49,index,index, "Cumple (S/N)",Format.SIZE(9))
    
    index += 2
    writer.range(2,5,index,index,   "", Format.SIZE(9), Format.BORDER)
    writer.range(6,9,index,index,   "", Format.SIZE(9), Format.BORDER)
    writer.range(10,13,index,index, "", Format.SIZE(9), Format.BORDER)
    writer.range(14,18,index,index, "", Format.SIZE(9), Format.BORDER)
    writer.range(19,23,index,index, "", Format.SIZE(9), Format.BORDER)
    writer.range(24,28,index,index, "", Format.SIZE(9), Format.BORDER)
    writer.range(29,33,index,index, "", Format.SIZE(9), Format.BORDER)
    writer.range(34,38,index,index, "", Format.SIZE(9), Format.BORDER)
    writer.range(39,42,index,index, "", Format.SIZE(9), Format.BORDER)
    writer.range(43,49,index,index, "", Format.SIZE(9), Format.BORDER)
    writer.cell (50,index,"",Format.BLEFT)

    
    curr_row = 16
    
    COL_WIDTH = [ 0.1 for i in range(0,50) ]
    
    annexUtils.set_column(worksheet,COL_WIDTH)
    annexUtils.set_row_dict(worksheet,ROW_DICT)
    
    workbook.close()
    



if __name__ == "__main__":
    generate('','')
