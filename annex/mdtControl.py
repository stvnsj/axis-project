import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
from utils import str_to_flt
import sys
import reader as rd
import re
import annexUtils
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
from annexUtils import offset_range
from openpyxl   import load_workbook
from control.mdt import MDTControl
import level

ROW_DICT = {}

def generate (
        input1 = "",
        input2 = "",
        output_file = "/home/jstvns/axis/Anexos Formulas/mdtcontrol.xlsx",
        tramo = "*",
        total_length = 1000,
) :
    
    workbook   = xlsxwriter.Workbook(output_file)
    worksheet  = workbook.add_worksheet("SHEET1")
    mdt_control = MDTControl(input1,input2)
    
    # worksheet.hide_gridlines(2) 
    # worksheet.set_portrait()
    # worksheet.set_page_view(2)
    # worksheet.set_paper(9)
    # worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)

    worksheet.hide_gridlines(2)  # 2 hides both the printed and visible gridlines
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(5)
    
    writer = Writer(workbook,worksheet)

    writer.range(2,19,2,6,"",Format.BORDER)
    writer.range(20,49,2,4,"SOMETHING ELSE",
                 Format.SIZE(12),Format.BOLD, Format.BTOP,Format.BRIGHT, Format.CENTER, Format.VCENTER)
    writer.range(20,49,5,6,"FORMULARIO N ####",
                 Format.SIZE(12),Format.BOLD, Format.BBOTTOM,Format.BRIGHT, Format.BOTTOM, Format.CENTER)

    
    
    writer.range(1,1,8,11,"",Format.BRIGHT)
    writer.range(50,50,8,11,"",Format.BLEFT)
    writer.range(2,49,7,7,"",Format.BBOTTOM)
    writer.range(2,49,12,12,"",Format.BTOP)
    
    writer.cell (2,8,"PROYECTO:",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.cell (2,9,"SECTOR:", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.cell (2,10,"TRAMO:", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.cell (2,11,"REALIZADO:",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.range(36,49,11,11,f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
    
    writer.cell(10,8, "NOMBRE DEL PROYECTO")
    writer.cell(10,9, "SECTOR DEL PROYECTO")
    writer.cell(10,10,"TRAMO DEL PROYECTO")
    writer.cell(10,11,"REALIZADO POR EQC")
    
    
    
    
    
    # writer.range(2,6,14,14,"Tramo N°:",Format.SIZE(9))
    # writer.range(7,11,14,14,tramo,Format.SIZE(9),Format.CENTER, Format.BBOTTOM) #PROGRAM
    # writer.range(13,17,14,14,"Dm. Inicio:" ,Format.SIZE(9))
    # writer.range(18,23,14,14,axis_ctrl.min_ctrl_dm ,Format.SIZE(9),Format.CENTER, Format.BBOTTOM) #PROGRAM
    # writer.range(25,28,14,14,"Dm. Fin:" ,Format.SIZE(9))
    # writer.range(29,33,14,14,axis_ctrl.max_ctrl_dm,Format.SIZE(9),Format.CENTER, Format.BBOTTOM) #PROGRAM
    # #writer.range(35,42,14,14,"Longitud del Tramo:",Format.SIZE(9))
    # #writer.range(43,49,14,14,axis_ctrl.ctrl_length,Format.SIZE(9),Format.CENTER, Format.BORDER) #PROGRAM
    
    
    # writer.range(40,45,14,14,"% muestral:", Format.SIZE(9))
    # writer.range(46,49,14,14,np.round(100*axis_ctrl.ctrl_length/total_length,1),Format.SIZE(9),Format.BORDER,Format.NUM1,Format.CENTER)

    # writer.range(34,45,16,16,"Tolerancia (m): ", Format.RIGHT,Format.SIZE(10))
    # writer.range(46,49,16,16,"0,10 m", Format.CENTER,Format.SIZE(10),Format.BORDER,Format.CENTER, Format.BORDER)
    
    index = 17
    ROW  = (index,index)
    
    
    dm_range = (2,7)
    n_range = offset_range(1,6,dm_range)
    side_range = offset_range(1,3,n_range)
    dist_range = offset_range(1,6,side_range)
    hc_range = offset_range(1,6,dist_range)
    hp_range = offset_range(1,6,hc_range)
    dif_range = offset_range(1,5,hp_range)
    absdif_range = offset_range(1,5,dif_range)
    good_range = offset_range(1,5,absdif_range)
    
    HEAD_FORMAT = (
        Format.VCENTER,
        Format.CENTER,
        Format.SIZE(10),
        Format.BORDER,
        Format.CENTER,
        Format.DOT
    )
    
    writer.range(*dm_range,index,index+1, "Perfil N°\n(Dm)", *HEAD_FORMAT)
    writer.range(*n_range,index,index+1,'Punto N°', *HEAD_FORMAT)
    writer.range(*side_range,index,index+1, "Lado", *HEAD_FORMAT)
    writer.range(*dist_range,index,index+1, "Distancia", *HEAD_FORMAT)
    writer.range(*hc_range,index,index+1, "Cota\nControl", *HEAD_FORMAT)
    writer.range(*hp_range,index,index+1,'Cota\nEstudio', *HEAD_FORMAT)
    writer.range(*dif_range,index,index+1,'Dif. (m)', *HEAD_FORMAT)
    writer.range(*absdif_range,index,index+1,'Dif. (m)\nValor Abs.', *HEAD_FORMAT)
    writer.range(*good_range,index,index+1,'Cumplr\n(S/N)', *HEAD_FORMAT)
    
    # writer.range(dif_range[0],good_range[1],index,index, 'Comparación',*HEAD_FORMAT)
    # writer.range(*dif_range,index+1,index+1,'Dif. Posición', *HEAD_FORMAT)
    # writer.range(*good_range,index+1,index+1,'Cumple (S/N)', *HEAD_FORMAT)
    
    # index+=2
    # DATA_FORMAT = (Format.CENTER,Format.SIZE(10),Format.BORDER,Format.CENTER, Format.BORDER,Format.NUM)
    
    
    # for point in axis_ctrl.point_list:
    #     COL = (index,index)
    #     writer.range(*dm_range,*COL,utils.str_to_flt(point.dm),*DATA_FORMAT)
    #     writer.range(*np_range,*COL,utils.str_to_flt(point.x_proj),*DATA_FORMAT)
    #     writer.range(*ep_range,*COL,utils.str_to_flt(point.y_proj),*DATA_FORMAT)
    #     writer.range(*nc_range,*COL,utils.str_to_flt(point.x_ctrl),*DATA_FORMAT)
    #     writer.range(*ec_range,*COL,utils.str_to_flt(point.y_ctrl),*DATA_FORMAT)
    #     writer.range(*dif_range,*COL,utils.str_to_flt(point.distance),*DATA_FORMAT)
    #     writer.range(*good_range,*COL,point.good,*DATA_FORMAT)
    #     index+=1
    
    # index += 2
    
    # writer.range(2,49,index,index,
    #              "Observaciones respecto de las mediciones",
    #              Format.SIZE(10),
    #              Format.BOLD)
    # index+=1
    # writer.range(2,49,index,index+1,"",
    #              Format.CENTER,Format.BORDER,Format.SIZE(10))

    # index+=3

    # writer.range(2,49,index,index,
    #              "Fotografías representativas de la materialización del estacado:",
    #              Format.SIZE(10),
    #              Format.BOLD,
    #              Format.CENTER)
    
    # index += 1
    # COLS = (index, index+8)
    
    # pic1_range = (2,16)
    # pic2_range = offset_range(2,17,pic1_range)
    # pic3_range = offset_range(2,16,pic2_range)
    
    # writer.range(*pic1_range,*COLS,"",Format.BORDER)
    # writer.range(*pic2_range,*COLS,"",Format.BORDER)
    # writer.range(*pic3_range,*COLS,"",Format.BORDER)
    
    # index += 9
    # writer.range(
    #     2,49,index,index,
    #     "Descripción general y observaciones respecto de la materialización del estacado:",
    #     Format.SIZE(10),
    #     Format.BOLD
    # )

    # index += 1
    # writer.range(
    #     2,49,index,index+3,
    #     "",Format.CENTER,Format.BORDER,Format.SIZE(10)
    # )

    # writer.cell(2,index+4,"",Format.BTOP)
    
    
    
    
    
    
    COL_WIDTH = [ 0.12 for i in range(0,50) ]
    
    annexUtils.set_column(worksheet,COL_WIDTH)
    workbook.close()
    



if __name__ == "__main__":
    generate(sys.argv[1],sys.argv[2])
