

import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
import annexUtils
from openpyxl import load_workbook


OFFSET   = 38
PAGEBREAKS = []

def generate (input_file='anexos/anteproyecto/anexo1.xlsx',output_file="test5.xlsx") :
    
    workbook = xlsxwriter.Workbook(output_file)
    worksheet = workbook.add_worksheet("PERFILES")
    writer = Writer(workbook,worksheet)
    
    COL_WIDTHS = [
        0.10, #A 
        0.30, #B
        0.70, #C
        0.25, #D
        0.39, #E
        0.18, #F
        0.19, #G
        0.19, #H
        0.25, #I
        0.22, #J
        0.30, #K
        0.25, #L
        0.15, #M
        0.42, #N
        0.15, #O
        0.20, #P
        0.20, #Q
        0.33, #R
        0.25, #S
        0.15, #T
        0.25, #U
        0.15, #V
        0.33, #W
        0.19, #X
        0.46, #Y
        0.25, #Z
        0.10, #AA
    ]
    
    ROW_DICT = {
        0 :0.1,  
        6 :0.1,  
        7 :0.18, 
        8 :0.18, 
        9 :0.19, 
        10:0.1,
        12:0.12,
        13:0.12
        
    }
    
    #worksheet.autofit()
    annexUtils.set_column(worksheet,COL_WIDTHS)
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
    
    # FIXED CONTENT
    writer.merge(f"B2:F6","",Format.BORDER)
    writer.merge(f"G2:Z3","PUNTOS DE LA RED DE REFERENCIA PRINCIPAL",
                 {"top":1, "right":1,"font_size":12,"bold":True,"align":"center","valign":"vcenter"})
    
    writer.merge(f"G4:Z6","FORMULARIO N° 2.903.3.F",
                 {"bottom":1,"right":1, "font_size":12,"bold":True,"align":"center","valign":"vcenter"})
    
    writer.write(f"AA2","",Format.LEFT)

    writer.merge(f"B7:Z7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"AA8:AA12","",{"left":1})
    writer.merge(f"B13:Z13", "",{"top":1})

    writer.merge(f"B8:C8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B9:C9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B10:C10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"B12:C12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"U12:Z12",f"FECHA: {annexUtils.curr_date()}",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
  
    

    
    
    wb = load_workbook(input_file)
    ws = wb.active
 
    scanner = annexUtils.Scanner(ws)
    
    t_rows = scanner.get_t_rows()
    i = 0
    
    # LOOP THE FOLLOWING CELLS 
    for r in t_rows :
        
        CELL_nombre = ws[f'C{r}'].value # DONE
        
        CELL_f      = ws[f'{scanner.GEO_S.column_letter}{r}'].value
        CELL_l      = ws[f'{scanner.GEO_W.column_letter}{r}'].value
        CELL_h      = ws[f'{scanner.ELIP.column_letter}{r}'].value

        POLY_NUM    = scanner.get_poligonal_num(r)
        
        CELL_X      = ws[f'{scanner.GEO_X.column_letter}{r}'].value
        CELL_Y      = ws[f'{scanner.GEO_Y.column_letter}{r}'].value
        CELL_Z      = ws[f'{scanner.GEO_Z.column_letter}{r}'].value
        
        CELL_N      = ws[f'{scanner.UTM_N.column_letter}{r}'].value
        CELL_E      = ws[f'{scanner.UTM_E.column_letter}{r}'].value
        
        CELL_altura = ws[f'{scanner.COTA_ORTO.column_letter}{r}'].value
        CELL_cota   = ws[f'{scanner.COTA_GEO.column_letter}{r}'].value
        
        CELL_NL     = ws[f'K{r}'].value
        CELL_EL     = ws[f'L{r}'].value
        
        CELL_MCL    = ws['H9'].value
        CELL_Ko     = ws['H12'].value
        
        
        
        writer.write(f"B{15 + i * OFFSET}","Identificación del Vértice",Format.BOLD, Format.ITALIC, Format.SIZE(10))
        writer.write(f"F{15 + i * OFFSET}", "Nombre:", Format.SIZE(10))
        writer.write(f"N{15 + i * OFFSET}", "Dm. Ref.:",Format.SIZE(10))
        
        writer.write(
            f'F{16 + i * OFFSET}',
            "Pertenece a Poligonal Nº:",
            Format.SIZE(11)
        )
        
        writer.merge(
            f'M{16 + i * OFFSET}:N{16 + i * OFFSET}',
            POLY_NUM,
            Format.BBOTTOM, Format.CENTER, Format.SIZE(11)
        )
        
        writer.write(
            f'P{16 + i * OFFSET}',
            "Tipo de Poligonal: (Ppal/Aux):",
            Format.SIZE(11)
        )
        
        writer.merge(
            f'X{16 + i * OFFSET}:Z{16 + i * OFFSET}',
            "",
            Format.BBOTTOM,Format.SIZE(11), Format.CENTER
        )
        
        writer.merge(
            f"I{15 + i * OFFSET}:L{15 + i * OFFSET}",
            CELL_nombre,
            Format.SIZE(11), Format.BOTTOM, Format.CENTER
        )
        
        writer.merge(
            f"N{25 + i * OFFSET}:P{25 + i * OFFSET}",
            CELL_altura,
            Format.RIGHT
        )
        
        writer.merge(
            f"X{25 + i * OFFSET}:Y{25 + i * OFFSET}",
            CELL_cota,
            Format.RIGHT
        )
        
        writer.merge(
            f'O{20 + i * OFFSET}:R{20 + i * OFFSET}',
            scanner.ZONA,
            
        )
        
        writer.merge(
            f'O{21 + i * OFFSET}:R{21 + i * OFFSET}',
            scanner.MC,
            
        )
        
        writer.merge(
            f"O{22 + i * OFFSET}:R{22 + i * OFFSET}",
            CELL_N,
            Format.RIGHT
        )
        
        writer.merge(
            f"O{23 + i * OFFSET}:R{23 + i * OFFSET}",
            CELL_E,
            Format.RIGHT
        )
        
        writer.merge(
            f"D{22 + i * OFFSET}:H{22 + i * OFFSET}",
            CELL_NL,
            Format.RIGHT
        )
        
        writer.merge(
            f"D{23 + i * OFFSET}:H{23 + i * OFFSET}",
            CELL_EL,
            Format.RIGHT
        )
        
        writer.merge(
            f"D{20 + i * OFFSET}:H{20 + i * OFFSET}",
            CELL_MCL,
            Format.LEFT
        )
        
        writer.merge(
            f"D{21 + i * OFFSET}:H{21 + i * OFFSET}",
            CELL_Ko,
            Format.LEFT
        )
        
        writer.merge(f"Q{15 + i * OFFSET}:T{15 + i * OFFSET}", "",Format.CENTER,Format.SIZE(10),Format.BOTTOM)
        writer.write(f"W{15 + i * OFFSET}","FECHA:",Format.SIZE(10))
        writer.merge(f"Y{15 + i * OFFSET}:Z{15 + i * OFFSET}", annexUtils.curr_date(1),Format.BOTTOM,Format.CENTER,Format.SIZE(10))
        
        writer.merge(f"N{18 + i * OFFSET}:S{19 + i * OFFSET}", "UTM",Format.SIZE(11),Format.CENTER,Format.VCENTER)
        writer.merge(f"D{18 + i * OFFSET}:H{19 + i * OFFSET}", "PTL",Format.SIZE(11),Format.CENTER,Format.VCENTER)
        
        writer.write(f"Q{25 + i * OFFSET}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"Z{25 + i * OFFSET}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"S{22 + i * OFFSET}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"S{23 + i * OFFSET}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"I{22 + i * OFFSET}","m",Format.SIZE(11),Format.LEFT)
        writer.write(f"I{23 + i * OFFSET}","m",Format.SIZE(11),Format.LEFT)
        
        writer.write(f"N{20 + i * OFFSET}","Huso:",Format.SIZE(11),Format.LEFT)
        writer.write(f"N{21 + i * OFFSET}","MC:",Format.SIZE(11),Format.LEFT)
        writer.write(f"N{22 + i * OFFSET}","N:",Format.SIZE(11),Format.LEFT)
        writer.write(f"N{23 + i * OFFSET}","E:",Format.SIZE(11),Format.LEFT)
        
        writer.write(f"C{20 + i * OFFSET}","MCL:",Format.SIZE(11),Format.LEFT)
        writer.write(f"C{21 + i * OFFSET}","Ko:",Format.SIZE(11),Format.LEFT)
        writer.write(f"C{22 + i * OFFSET}","NL:",Format.SIZE(11),Format.LEFT)
        writer.write(f"C{23 + i * OFFSET}","EL:",Format.SIZE(11),Format.LEFT)
        
        writer.write(f"F{25 + i * OFFSET}","Altura (n.m.m. modelada):",Format.SIZE(11))
        writer.write(f"S{25 + i * OFFSET}","Cota (nivelada):",Format.SIZE(11),Format.BOLD)
        
        writer.merge(f"B{27 + i * OFFSET}:K{39 + i * OFFSET}","Fotografía\nPanorámica",Format.BORDER,Format.CENTER,Format.VCENTER)
        writer.merge(f"M{27 + i * OFFSET}:Z{39 + i * OFFSET}","Vista\nAérea",Format.BORDER,Format.CENTER,Format.VCENTER)
        writer.merge(f"B{41 + i * OFFSET}:F{48 + i * OFFSET}","Fotografía\nDetalle",Format.BORDER,Format.CENTER,Format.VCENTER)
        
        
        writer.merge(f"H{41 + i * OFFSET}:Z{41 + i * OFFSET}","Descripción",Format.BOTTOM,Format.LEFT,Format.SIZE(10))
        
        writer.write(f'I{43 + i * OFFSET}', "Materialidad:",Format.SIZE(9))
        writer.write(f'I{44 + i * OFFSET}', "Dimensiones:",Format.SIZE(9))
        writer.write(f'I{45 + i * OFFSET}', "Distancia a :",Format.SIZE(9))
        
        writer.write(f'L{43 + i * OFFSET}', "MONOLITO DE HORMIGÓN, PINTADO DE AMARILLO", Format.SIZE(9))
        writer.write(f'L{44 + i * OFFSET}', "D: 15 cm. FIERRO ESTRIADO 12 mm.", Format.SIZE(9))
        
        writer.merge(f"G{42 + i * OFFSET}:G{48 + i * OFFSET}","",Format.BRIGHT)
        writer.merge(f"AA{42 + i * OFFSET}:AA{48 + i * OFFSET}","",Format.BLEFT)
        writer.merge(f"H{49 + i * OFFSET}:Z{49 + i * OFFSET}","",Format.TOP)
        writer.merge(f"A{50 + i * OFFSET}:AA{50 + i * OFFSET}","",{})
        
        ROW_DICT.update({15 + i * OFFSET : 0.25})
        ROW_DICT.update({16 + i * OFFSET : 0.35})
        ROW_DICT.update({key:0.18 for key in range(27 + i*OFFSET , 39 + i*OFFSET)})  
        
        PAGEBREAKS.append(50 + i * OFFSET)
        
        i += 1
    
    worksheet.set_h_pagebreaks(PAGEBREAKS)
    annexUtils.set_row_dict(worksheet,ROW_DICT)
    
    # formatter.set_rows({0:2,1:2,2:2, 4:2})
    workbook.close()




if __name__ == "__main__":
    
    generate()
    
    #f1 = sys.argv[1]
    #f2 = sys.argv[2]
    
    #reader = rd.Reader (f1, "", f2)
    #matrix, labels, om, ol, heights = reader.getData()
    #model = mdl.Model(heights,matrix,labels, om, ol)
    
    #trans (model)

