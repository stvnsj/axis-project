import xlsxwriter
import openpyxl
import numpy as np
import model as mdl
import utils
import sys
import reader as rd
import re
import annexUtils
from annexUtils import Format 
from annexUtils import Writer
from annexUtils import Formatter
from openpyxl   import load_workbook
import level


def generate (
        input1 = "",
        input2 = "",
        output_file = "testlongi.xlsx") :
    cir = level.parser(input1, input2)
    TABLE = cir.get_report_long()
    print(TABLE)
    
    workbook   = xlsxwriter.Workbook(output_file)
    worksheet  = workbook.add_worksheet()
    
    worksheet.hide_gridlines(2)
    worksheet.set_portrait()
    worksheet.set_page_view(2)
    worksheet.set_paper(9)
    worksheet.set_margins(left=0.71, right=0.71, top=0.95, bottom=0.75)
    
    writer = Writer(workbook,worksheet)
    
    # FIXED CONTENT
    writer.merge(f"B2:C6","",Format.BORDER)
    
    writer.merge(
        f"D2:H5","NIVELACIÓN LONGITUDINAL DEL EJE ESTACADO",
        Format.BTOP,Format.BRIGHT,Format.SIZE(12),Format.BOLD,
        Format.CENTER, Format.VCENTER
    )
    
    writer.merge(
        f"D6:H6",
        "FORMULARIO N° 2.5.3",
        Format.BBOTTOM,Format.BRIGHT,Format.SIZE(12),
        Format.BOLD,Format.CENTER
    )
    
    
    writer.write(f"I2","",Format.BLEFT)
    writer.merge(f"B7:H7","",{"bottom":1})
    writer.merge(f"A8:A12","",{"right":1})
    writer.merge(f"I8:I12","",{"left":1})
    writer.merge(f"B13:H13","",{"top":1})
    
    writer.write(f"B8","PROYECTO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B9","SECTOR", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B10","TRAMO", Format.SIZE(10),Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.write(f"B12","REALIZADO",Format.SIZE(10), Format.BOLD, Format.LEFT, Format.VCENTER)
    writer.merge(f"G12:H12","FECHA:",Format.SIZE(10),Format.RIGHT,Format.VCENTER)
  
    
    # wb = load_workbook(input_file)
    # ws = wb.active
    
    # max_row = ws.max_row
    # first_row = True
    # curr_row = 18
    # DESDE = 1
    # HASTA = 1
    # BREAK = False
 
    # POINT_A = ""
    # POINT_B = ""
    # DELTA_A = 0.0
    # DELTA_B = 0.0
    # ERR     = 0.0
    # DELTA_MEAN = 0.0
    # HEIGHT = 0.0
    
    
    COL_WIDTH = [
        0.13, # A
        0.40, # B
        0.40, # C
        0.40, # D
        0.40, # E
        0.40, # F
        0.40, # G
        0.40, # H
        0.13, # I   
    ]
    annexUtils.set_column(worksheet,COL_WIDTH)
    
    workbook.close()
    






if __name__ == "__main__":
    f1 = sys.argv[1]
    f2 = sys.argv[2]
    generate(f1, f2)
